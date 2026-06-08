from __future__ import annotations

import io
import html
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def to_excel_bytes(sheets_dict: dict[str, pd.DataFrame]) -> bytes:
    """Pack a {sheet_name: DataFrame} dict into XLSX bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets_dict.items():
            safe_name = str(name)[:31]   # Excel sheet name max 31 chars
            if df is None or (isinstance(df, pd.DataFrame) and df.empty):
                pd.DataFrame({"(empty)": []}).to_excel(writer, sheet_name=safe_name, index=False)
            else:
                df.to_excel(writer, sheet_name=safe_name, index=False)
    output.seek(0)
    return output.getvalue()


def to_styled_duetto_pivot_excel_bytes(pivot_df: pd.DataFrame) -> bytes:
    """Build a one-sheet styled XLSX for the Duetto Pivot - by Stay Month table."""
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    if pivot_df is None or pivot_df.empty:
        _write_styled_pivot_sheet(wb, "Duetto Pivot", pd.DataFrame())
    else:
        _write_styled_pivot_sheet(wb, "Duetto Pivot", pivot_df)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def to_duetto_pivot_svg_bytes(pivot_df: pd.DataFrame, title: str = "Duetto Pivot - by Stay Month") -> bytes:
    """Render the full Duetto pivot as one dependency-free SVG image."""
    if pivot_df is None or pivot_df.empty:
        pivot_df = pd.DataFrame({"Message": ["No pivot data for selected filters."]})

    df = pivot_df.reset_index(drop=True)
    columns = list(df.columns)
    rows = df.to_dict("records")

    widths = _image_column_widths(df)
    header_h = 32
    row_h = 28
    title_h = 46
    margin = 14
    canvas_w = sum(widths) + margin * 2
    canvas_h = title_h + header_h + row_h * len(rows) + margin * 2

    parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{canvas_w}" height="{canvas_h}" '
            f'viewBox="0 0 {canvas_w} {canvas_h}">'
        ),
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        f'<rect x="0" y="0" width="{canvas_w}" height="{title_h + margin}" fill="#F8FAFC"/>',
        (
            f'<text x="{margin}" y="42" fill="#111827" font-family="Arial, sans-serif" '
            f'font-size="24" font-weight="700">{html.escape(title)}</text>'
        ),
    ]

    y = title_h + margin
    x = margin
    for col, width in zip(columns, widths):
        parts.append(f'<rect x="{x}" y="{y}" width="{width}" height="{header_h}" fill="#111827"/>')
        parts.append(_svg_text(str(col), x + 5, y + 22, width - 10, "#FFFFFF", bold=True))
        x += width

    y += header_h
    for row in rows:
        metric = str(row.get("Metric", ""))
        x = margin
        for col, width in zip(columns, widths):
            value = _format_png_value(row.get(col), col)
            fill, text_color, bold = _image_cell_colors(col, metric, row.get(col))
            parts.append(f'<rect x="{x}" y="{y}" width="{width}" height="{row_h}" fill="{fill}"/>')
            parts.append(
                f'<line x1="{x}" y1="{y + row_h}" x2="{x + width}" y2="{y + row_h}" stroke="#E5E7EB"/>'
            )
            parts.append(f'<line x1="{x + width}" y1="{y}" x2="{x + width}" y2="{y + row_h}" stroke="#E5E7EB"/>')
            if col in {"Hotel", "Stay Month", "Metric", "Message"}:
                parts.append(_svg_text(value, x + 5, y + 20, width - 10, text_color, bold=bold))
            else:
                parts.append(
                    _svg_text(value, x + width - 5, y + 20, width - 10, text_color, bold=bold, anchor="end")
                )
            x += width
        y += row_h

    parts.append("</svg>")
    return "\n".join(parts).encode("utf-8")


def _safe_sheet_name(name: str) -> str:
    safe = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip()
    return (safe or "Sheet")[:31]


def _write_styled_pivot_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    _apply_one_page_export_setup(ws)

    if df is None or df.empty:
        ws.append(["No pivot data for selected filters."])
        ws["A1"].font = Font(bold=True, color="6B7280")
        ws.column_dimensions["A"].width = 34
        return

    columns = list(df.columns)
    ws.append(columns)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="D1D5DB"))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for _, row in df.iterrows():
        ws.append([row.get(col) for col in columns])

    ws.freeze_panes = "D2" if "Hotel" in columns else "C2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx in range(2, ws.max_row + 1):
        values = {col: ws.cell(row=row_idx, column=idx + 1).value for idx, col in enumerate(columns)}
        metric = str(values.get("Metric", ""))
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            _apply_pivot_cell_style(cell, col, metric, values)

    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 24)

    ws.sheet_view.showGridLines = False


def _apply_one_page_export_setup(ws) -> None:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25


def _apply_pivot_cell_style(cell, col: str, metric: str, row_values: dict[str, object]) -> None:
    base_fill = None
    if metric == "Rev":
        base_fill = "FFFBF5"
        cell.font = Font(bold=True if col in {"Today", "Duetto"} else False, color="111827")
    elif metric == "Occ":
        base_fill = "F9FAFB"

    if base_fill:
        cell.fill = PatternFill("solid", fgColor=base_fill)

    if col in {"Hotel", "Stay Month", "Metric"}:
        cell.font = Font(bold=(col == "Metric"), color="111827")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="right", vertical="center")

    if col in {"Today", "STLY", "ST2Y", "ST3Y", "Duetto", "Budget", "Final LY", "Final 2Y", "Final 3Y"}:
        cell.number_format = '#,##0.00'
    if " VS " in str(col):
        cell.number_format = '+0.0"%"'

    if col == "Today":
        _set_cell_colors(cell, "DBEAFE" if metric == "Rev" else "EFF6FF", "1E40AF", True)
    elif col == "Duetto":
        _set_cell_colors(cell, "DCFCE7" if metric == "Rev" else "F0FDF4", "15803D", True)
    elif col == "Budget":
        _set_cell_colors(cell, "FEFCE8", "713F12", metric == "Rev")

    if " VS " in str(col):
        try:
            value = float(row_values.get(col))
        except (TypeError, ValueError):
            value = None
        if value is not None:
            if value > 0:
                _set_cell_colors(cell, "BBF7D0", "166534", True)
            elif value < 0:
                _set_cell_colors(cell, "FECACA", "991B1B", True)
            else:
                _set_cell_colors(cell, "FEF9C3", "92400E", True)

    cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))


def _set_cell_colors(cell, fill: str, font_color: str, bold: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=bold, color=font_color)


def _image_column_widths(df: pd.DataFrame) -> list[int]:
    widths = []
    for col in df.columns:
        values = [str(col)] + [_format_png_value(v, str(col)) for v in df[col].head(80)]
        measured = max(len(v) for v in values) * 7 + 18
        if col == "Hotel":
            widths.append(min(max(measured, 130), 190))
        elif col == "Stay Month":
            widths.append(min(max(measured, 92), 115))
        elif col == "Metric":
            widths.append(58)
        elif " VS " in str(col):
            widths.append(min(max(measured, 78), 102))
        else:
            widths.append(min(max(measured, 78), 100))
    return widths


def _format_png_value(value, col: str) -> str:
    if pd.isna(value):
        return "-"
    if " VS " in str(col):
        try:
            return f"{float(value):+.1f}%"
        except (TypeError, ValueError):
            return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.2f}"
    return str(value)


def _image_cell_colors(col: str, metric: str, value) -> tuple[str, str, bool]:
    fill = "#FFFFFF"
    text = "#111827"
    bold = False

    if metric == "Rev":
        fill = "#FFFBF5"
    elif metric == "Occ":
        fill = "#F9FAFB"

    if col == "Today":
        return ("#DBEAFE" if metric == "Rev" else "#EFF6FF", "#1E40AF", True)
    if col == "Duetto":
        return ("#DCFCE7" if metric == "Rev" else "#F0FDF4", "#15803D", True)
    if col == "Budget":
        return ("#FEFCE8", "#713F12", metric == "Rev")

    if " VS " in str(col):
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            numeric = None
        if numeric is not None:
            if numeric > 0:
                return "#BBF7D0", "#166534", True
            if numeric < 0:
                return "#FECACA", "#991B1B", True
            return "#FEF9C3", "#92400E", True

    return fill, text, bold


def _svg_text(
    text: str,
    x: int,
    y: int,
    max_width: int,
    fill: str,
    bold: bool = False,
    anchor: str = "start",
) -> str:
    max_chars = max(3, max_width // 7)
    clipped = text if len(text) <= max_chars else text[: max_chars - 3] + "..."
    weight = "700" if bold else "400"
    return (
        f'<text x="{x}" y="{y}" fill="{fill}" font-family="Arial, sans-serif" '
        f'font-size="13" font-weight="{weight}" text-anchor="{anchor}">{html.escape(clipped)}</text>'
    )
