from __future__ import annotations

import io
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook

from src.services.excel_export import to_duetto_pivot_svg_bytes, to_styled_duetto_pivot_excel_bytes


def _fill(cell) -> str:
    return str(cell.fill.fgColor.rgb or cell.fill.fgColor.indexed)


def test_styled_duetto_pivot_excel_keeps_variance_colors():
    df = pd.DataFrame(
        [
            {
                "Hotel": "G5 Test Hotel",
                "Stay Month": "Jun 2026",
                "Metric": "Rev",
                "Today": 120,
                "Budget": 100,
                "Duetto": 130,
                "Today VS BUD": 20,
                "Duetto VS BUD": -10,
            }
        ]
    )

    wb = load_workbook(io.BytesIO(to_styled_duetto_pivot_excel_bytes(df)))
    ws = wb["Duetto Pivot"]

    assert wb.sheetnames == ["Duetto Pivot"]
    assert ws.freeze_panes == "D2"
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 1
    assert _fill(ws["G2"]).endswith("BBF7D0")
    assert _fill(ws["H2"]).endswith("FECACA")
    assert ws["G2"].number_format == '+0.0"%"'


def test_styled_duetto_pivot_excel_writes_empty_state():
    wb = load_workbook(io.BytesIO(to_styled_duetto_pivot_excel_bytes(pd.DataFrame())))

    assert wb.sheetnames == ["Duetto Pivot"]
    assert wb["Duetto Pivot"]["A1"].value == "No pivot data for selected filters."


def test_duetto_pivot_svg_export_writes_single_image_canvas():
    df = pd.DataFrame(
        [
            {
                "Hotel": "G5 Test Hotel",
                "Stay Month": "Jun 2026",
                "Metric": "Rev",
                "Today": 120,
                "Budget": 100,
                "Duetto": 130,
                "Today VS BUD": 20,
                "Duetto VS BUD": -10,
            }
        ]
    )

    svg = to_duetto_pivot_svg_bytes(df)
    root = ET.fromstring(svg)

    assert root.tag.endswith("svg")
    assert 700 <= int(root.attrib["width"]) <= 900
    assert int(root.attrib["height"]) > 80
    assert b"G5 Test Hotel" in svg
    assert b"#BBF7D0" in svg
    assert b"#FECACA" in svg
